"""Read Excel inputs, compute result rows, query SQLite, and export xlsx."""

from __future__ import annotations

import io
import re
import shutil
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

from config import settings as S

DATE_RE = re.compile(S.DATE_PATTERN)


def normalize_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    return text if DATE_RE.match(text) else None


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def load_sheet(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    try:
        rows = ws.iter_rows(values_only=True)
        raw_header = next(rows)
    except StopIteration:
        wb.close()
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in raw_header]
    result: list[dict[str, Any]] = []
    for raw_row in rows:
        item: dict[str, Any] = {}
        for index, name in enumerate(header):
            if not name:
                continue
            item[name] = raw_row[index] if index < len(raw_row) else None
        if any(value is not None for value in item.values()):
            result.append(item)
    wb.close()
    return result


def validate_headers(sheet_rows: list[dict[str, Any]], required: set[str]) -> list[str]:
    if not sheet_rows:
        return sorted(required)
    return sorted(required - set(sheet_rows[0].keys()))


def aggregate_tongshi(tongshi_rows: list[dict[str, Any]]) -> dict[tuple[str, str, str], dict[str, float | None]]:
    agg: dict[tuple[str, str, str], dict[str, float | None]] = {}
    for row in tongshi_rows:
        day = normalize_date(row.get("日期"))
        mode = row.get("模式")
        xuebu = row.get("学部")
        if not day or mode is None or xuebu is None:
            continue
        mode = str(mode).strip()
        xuebu = str(xuebu).strip()
        if not mode or not xuebu:
            continue
        bucket = agg.setdefault((day, mode, xuebu), {"hd": 0.0, "ex": 0.0, "ai": 0.0, "rate": None})
        bucket["hd"] = float(bucket["hd"] or 0.0) + _to_float(row.get("话单分钟数"))
        examples = _to_float(row.get("例子数"))
        bucket["ex"] = float(bucket["ex"] or 0.0) + examples
        bucket["ai"] = float(bucket["ai"] or 0.0) + _to_float(row.get("AI接通数"))

    for bucket in agg.values():
        examples = float(bucket.get("ex") or 0.0)
        ai_connected = float(bucket.get("ai") or 0.0)
        if ai_connected:
            bucket["rate"] = examples / ai_connected
    return agg


def labor_cost(mode: str, efficiency: float) -> float | None:
    if mode in S.LINE_ONLY_COST_MODES:
        return 0.0
    if mode in S.LABOR_COST_RULES:
        return S.LABOR_COST_RULES[mode]
    if mode == "大神" and efficiency:
        return S.DSHEN_LABOR_NUMERATOR / efficiency + S.DSHEN_LABOR_EXTRA
    return None


def compute_result(
    tongshi_rows: list[dict[str, Any]], zhuanhua_rows: list[dict[str, Any]]
) -> tuple[list[tuple[Any, ...]], list[tuple[str, str, str]]]:
    agg = aggregate_tongshi(tongshi_rows)
    by_mode: dict[tuple[str, str], dict[str, dict[str, float | None]]] = {}
    for (day, mode, xuebu), values in agg.items():
        by_mode.setdefault((day, mode), {})[xuebu] = values

    rows: list[tuple[Any, ...]] = []
    skipped: list[tuple[str, str, str]] = []
    for row in zhuanhua_rows:
        day = normalize_date(row.get("日期"))
        mode_raw = row.get("流转模式")
        if not day:
            skipped.append((str(row.get("日期")), str(mode_raw), "日期非标准格式（疑似加密痕迹行）"))
            continue
        mode = str(mode_raw).strip() if mode_raw is not None else ""
        if not mode:
            skipped.append((day, mode, "流转模式为空"))
            continue
        volume = _to_float(row.get("单量"))
        attendance = _to_float(row.get("出勤"))
        if attendance == 0.0:
            skipped.append((day, mode, "出勤为空或为 0"))
            continue
        efficiency = volume / attendance
        if efficiency == 0.0:
            skipped.append((day, mode, "单量为空或为 0"))
            continue
        human_cost = labor_cost(mode, efficiency)
        if human_cost is None:
            skipped.append((day, mode, "未知流转模式，无成本规则"))
            continue
        xuebu_map = by_mode.get((day, mode))
        if not xuebu_map:
            skipped.append((day, mode, "tongshi 无该组合明细（无可聚合的学部）"))
            continue
        for xuebu, values in xuebu_map.items():
            examples = float(values["ex"] or 0.0)
            if examples == 0.0:
                skipped.append((day, mode, f"学部 {xuebu} 例子数为 0"))
                continue
            line_cost = round(S.LINE_COST_UNIT * float(values["hd"] or 0.0) / examples, S.COST_DECIMALS)
            settlement_cost = round(human_cost + line_cost, S.COST_DECIMALS)
            rate = values.get("rate")
            rows.append(
                (
                    day,
                    mode,
                    xuebu,
                    efficiency,
                    line_cost,
                    settlement_cost,
                    rate,
                    examples,
                    attendance,
                    volume,
                    values.get("ai"),
                )
            )
    return rows, skipped


def init_db() -> None:
    S.DATA_DIR.mkdir(parents=True, exist_ok=True)
    if S.DB_PATH != S.SEED_DB_PATH and not S.DB_PATH.exists() and S.SEED_DB_PATH.exists():
        shutil.copy2(S.SEED_DB_PATH, S.DB_PATH)
    conn = sqlite3.connect(S.DB_PATH)
    conn.execute(
        """CREATE TABLE IF NOT EXISTS result(
            日期 TEXT NOT NULL,
            流转模式 TEXT NOT NULL,
            学部 TEXT NOT NULL,
            人效 REAL,
            线路成本 REAL,
            单例子结算成本 REAL,
            接通转化率 REAL,
            单量 REAL,
            出勤 REAL,
            人效单量 REAL,
            AI接通数 REAL,
            PRIMARY KEY(日期, 流转模式, 学部))"""
    )
    columns = [row[1] for row in conn.execute("PRAGMA table_info(result)").fetchall()]
    for column in ("单量", "出勤", "人效单量", "AI接通数"):
        if column not in columns:
            conn.execute(f"ALTER TABLE result ADD COLUMN {column} REAL")
    conn.commit()
    conn.close()


def upsert_rows(rows: list[tuple[Any, ...]]) -> None:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    try:
        conn.execute("DELETE FROM result")
        conn.executemany(
            "INSERT INTO result(日期, 流转模式, 学部, 人效, 线路成本, 单例子结算成本, 接通转化率, 单量, 出勤, 人效单量, AI接通数) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _xuebu_in_sql() -> str:
    return "学部 IN (" + ",".join("?" for _ in S.XUBU_WHITELIST) + ")"


def _rank_sql(column: str, order: dict[str, int]) -> str:
    cases = " ".join(f"WHEN '{key}' THEN {value}" for key, value in order.items())
    return f"CASE {column} {cases} ELSE 9 END"


def fetch_filtered(
    view: str = "latest",
    start: str | None = None,
    end: str | None = None,
    mode: str | list[str] | tuple[str, ...] | None = None,
    xuebu: str | list[str] | tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    conn.row_factory = sqlite3.Row
    where = [_xuebu_in_sql()]
    params: list[Any] = list(S.XUBU_WHITELIST)

    if view == "latest":
        where.append("日期 = (SELECT MAX(日期) FROM result WHERE " + _xuebu_in_sql() + ")")
        params += list(S.XUBU_WHITELIST)
    elif view == "custom":
        if start and DATE_RE.match(start):
            where.append("日期 >= ?")
            params.append(start)
        if end and DATE_RE.match(end):
            where.append("日期 <= ?")
            params.append(end)
    elif view in {"7d", "30d"}:
        days = 7 if view == "7d" else 30
        since = (datetime.now() - timedelta(days=days - 1)).strftime("%Y-%m-%d")
        where.append("日期 >= ?")
        params.append(since)
    elif view and view != "all" and DATE_RE.match(view):
        where.append("日期 = ?")
        params.append(view)

    modes = _normalize_modes(mode)
    if modes:
        where.append("流转模式 IN (" + ",".join("?" for _ in modes) + ")")
        params.extend(modes)
    xuebus = _normalize_xuebus(xuebu)
    if xuebus:
        where.append("学部 IN (" + ",".join("?" for _ in xuebus) + ")")
        params.extend(xuebus)

    sql = (
        "SELECT 日期, 流转模式, 学部, 人效, 线路成本, 单例子结算成本, 接通转化率, 单量, 出勤, 人效单量, AI接通数 "
        "FROM result WHERE "
        + " AND ".join(where)
        + " ORDER BY 日期 DESC, "
        + _rank_sql("流转模式", S.MODE_ORDER)
        + ", "
        + _rank_sql("学部", S.XUBU_ORDER)
        + ", 流转模式 ASC, 学部 ASC"
    )
    rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
    conn.close()
    return rows


def _normalize_modes(mode: str | list[str] | tuple[str, ...] | None) -> list[str]:
    if mode is None:
        return []
    if isinstance(mode, str):
        values = [mode]
    else:
        values = list(mode)
    return [item for item in (str(value).strip() for value in values) if item]


def _normalize_xuebus(xuebu: str | list[str] | tuple[str, ...] | None) -> list[str]:
    if xuebu is None:
        return []
    if isinstance(xuebu, str):
        values = [xuebu]
    else:
        values = list(xuebu)
    return [item for item in (str(value).strip() for value in values) if item in S.XUBU_ORDER]


def _apply_date_filter(where: list[str], params: list[Any], view: str, start: str | None, end: str | None) -> None:
    if view == "latest":
        where.append("日期 = (SELECT MAX(日期) FROM result WHERE " + _xuebu_in_sql() + ")")
        params += list(S.XUBU_WHITELIST)
    elif view == "custom":
        if start and DATE_RE.match(start):
            where.append("日期 >= ?")
            params.append(start)
        if end and DATE_RE.match(end):
            where.append("日期 <= ?")
            params.append(end)
    elif view in {"7d", "30d"}:
        days = 7 if view == "7d" else 30
        since = (datetime.now() - timedelta(days=days - 1)).strftime("%Y-%m-%d")
        where.append("日期 >= ?")
        params.append(since)
    elif view and view != "all" and DATE_RE.match(view):
        where.append("日期 = ?")
        params.append(view)


def fetch_cost_trend(
    view: str = "latest",
    start: str | None = None,
    end: str | None = None,
    mode: str | list[str] | tuple[str, ...] | None = None,
    xuebu: str | list[str] | tuple[str, ...] | None = None,
) -> list[dict[str, Any]]:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    conn.row_factory = sqlite3.Row
    where = [_xuebu_in_sql(), "COALESCE(单量, 0) > 0"]
    params: list[Any] = list(S.XUBU_WHITELIST)

    _apply_date_filter(where, params, view, start, end)

    modes = _normalize_modes(mode)
    if modes:
        where.append("流转模式 IN (" + ",".join("?" for _ in modes) + ")")
        params.extend(modes)
    xuebus = _normalize_xuebus(xuebu)
    if xuebus:
        where.append("学部 IN (" + ",".join("?" for _ in xuebus) + ")")
        params.extend(xuebus)

    sql = (
        "SELECT 日期, ROUND(SUM(单例子结算成本 * 单量) / SUM(单量), ?) AS 聚合单例子结算成本, SUM(单量) AS 总单量 "
        "FROM result WHERE "
        + " AND ".join(where)
        + " GROUP BY 日期 ORDER BY 日期 ASC"
    )
    rows = [dict(row) for row in conn.execute(sql, [S.COST_DECIMALS] + params).fetchall()]
    conn.close()
    return rows


def distinct_modes() -> list[str]:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    rows = conn.execute(
        "SELECT DISTINCT 流转模式 FROM result WHERE " + _xuebu_in_sql() + " ORDER BY 流转模式",
        list(S.XUBU_WHITELIST),
    ).fetchall()
    conn.close()
    return [row[0] for row in rows]


def distinct_dates() -> list[str]:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    rows = conn.execute(
        "SELECT DISTINCT 日期 FROM result WHERE " + _xuebu_in_sql() + " ORDER BY 日期 DESC",
        list(S.XUBU_WHITELIST),
    ).fetchall()
    conn.close()
    return [row[0] for row in rows]


def latest_date() -> str | None:
    init_db()
    conn = sqlite3.connect(S.DB_PATH)
    row = conn.execute("SELECT MAX(日期) FROM result WHERE " + _xuebu_in_sql(), list(S.XUBU_WHITELIST)).fetchone()
    conn.close()
    return row[0] if row else None


def generate_result_xlsx(rows: list[dict[str, Any] | tuple[Any, ...]]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(S.RESULT_HEADERS)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in rows:
        if isinstance(row, dict):
            ws.append([_export_value(key, row.get(key)) for key in S.RESULT_HEADERS])
        else:
            ws.append([_export_value(key, value) for key, value in zip(S.RESULT_HEADERS, row)])
    for col, width in zip("ABCDEFGH", [14, 16, 10, 10, 10, 14, 18, 14]):
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A1:H{max(ws.max_row, 1)}"
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _export_value(key: str, value: Any) -> Any:
    if key == "单量" and value is not None:
        return int(round(float(value)))
    if key in {"人效", "线路成本", "单例子结算成本"} and value is not None:
        return round(float(value), 1)
    return value


def save_upload(fileobj: Any, kind: str) -> Path:
    S.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = S.UPLOAD_DIR / f"{stamp}_{kind}.xlsx"
    fileobj.save(str(path))
    return path
