import tempfile
import unittest
from pathlib import Path

import openpyxl

import pipeline
from config import settings as S


class TestPipeline(unittest.TestCase):
    def test_compute_result_uses_existing_cost_rules(self):
        tongshi_rows = [
            {"日期": "2026-08-01", "模式": "大神", "学部": "小学", "例子数": 10, "话单分钟数": 100},
            {"日期": "2026-08-01", "模式": "爆量算法池", "学部": "高中", "例子数": 20, "话单分钟数": 80},
        ]
        zhuanhua_rows = [
            {"日期": "2026-08-01", "流转模式": "大神", "人效": 5, "接通转化率": 0.01},
            {"日期": "2026-08-01", "流转模式": "爆量算法池", "人效": 8, "接通转化率": 0.02},
        ]

        rows, skipped = pipeline.compute_result(tongshi_rows, zhuanhua_rows)

        self.assertEqual([], skipped)
        by_key = {(r[1], r[2]): r for r in rows}
        self.assertAlmostEqual(by_key[("大神", "小学")][4], 0.85)
        self.assertAlmostEqual(by_key[("大神", "小学")][5], 86.85)
        self.assertAlmostEqual(by_key[("爆量算法池", "高中")][5], 40.34)

    def test_fetch_filtered_latest_and_custom_range(self):
        original_db = S.DB_PATH
        with tempfile.TemporaryDirectory() as tmp:
            S.DB_PATH = Path(tmp) / "test.db"
            pipeline.upsert_rows(
                [
                    ("2026-08-01", "大神", "小学", 4.0, 10.0, 100.0, 0.01, 10.0),
                    ("2026-08-02", "9.9池", "初中", 2.0, 20.0, 110.0, 0.02, 20.0),
                    ("2026-08-03", "爆量算法池", "其他", 5.0, 30.0, 120.0, 0.03, 30.0),
                ]
            )

            latest = pipeline.fetch_filtered()
            custom = pipeline.fetch_filtered(view="custom", start="2026-08-01", end="2026-08-02")

            self.assertEqual(["2026-08-02"], [row["日期"] for row in latest])
            self.assertEqual(["2026-08-02", "2026-08-01"], [row["日期"] for row in custom])
        S.DB_PATH = original_db

    def test_fetch_cost_trend_weights_selected_modes_by_examples(self):
        original_db = S.DB_PATH
        with tempfile.TemporaryDirectory() as tmp:
            S.DB_PATH = Path(tmp) / "test.db"
            pipeline.upsert_rows(
                [
                    ("2026-08-01", "大神", "小学", 4.0, 10.0, 100.0, 0.01, 10.0),
                    ("2026-08-01", "爆量算法池", "高中", 5.0, 20.0, 40.0, 0.02, 30.0),
                    ("2026-08-01", "9.9池", "初中", 2.0, 30.0, 160.0, 0.03, 60.0),
                    ("2026-08-02", "大神", "小学", 4.0, 10.0, 120.0, 0.01, 20.0),
                    ("2026-08-02", "爆量算法池", "高中", 5.0, 20.0, 60.0, 0.02, 20.0),
                ]
            )

            trend = pipeline.fetch_cost_trend(view="all", mode=["大神", "爆量算法池"])

            self.assertEqual(["2026-08-01", "2026-08-02"], [row["日期"] for row in trend])
            self.assertAlmostEqual(trend[0]["聚合单例子结算成本"], 55.0)
            self.assertAlmostEqual(trend[0]["总单量"], 40.0)
            self.assertAlmostEqual(trend[1]["聚合单例子结算成本"], 90.0)
        S.DB_PATH = original_db

    def test_generate_result_xlsx_headers(self):
        buf = pipeline.generate_result_xlsx(
            [{"日期": "2026-08-01", "流转模式": "大神", "学部": "小学", "人效": 4.12, "线路成本": 1.23, "单例子结算成本": 2.34, "接通转化率": 0.01}]
        )
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb.active
        self.assertEqual(S.RESULT_HEADERS, [ws.cell(1, col).value for col in range(1, 8)])
        self.assertEqual("A1:G2", ws.auto_filter.ref)
        self.assertEqual(4.1, ws.cell(2, 4).value)
        self.assertEqual(1.2, ws.cell(2, 5).value)
        self.assertEqual(2.3, ws.cell(2, 6).value)


if __name__ == "__main__":
    unittest.main(verbosity=2)
